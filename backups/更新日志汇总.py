#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
可靠性测试报告自动化工具 - 更新日志汇总生成脚本
生成Excel格式的版本更新记录
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

def create_changelog_excel():
    """创建更新日志Excel文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "开发全周期记录"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 80
    
    # 大标题样式
    title_font = Font(name='微软雅黑', size=16, bold=True)
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    # 列标题样式
    header_font = Font(name='微软雅黑', size=11, bold=True)
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # 内容样式
    content_font = Font(name='微软雅黑', size=10)
    content_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    date_alignment = Alignment(horizontal='center', vertical='center')
    version_alignment = Alignment(horizontal='center', vertical='center')
    title_content_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入大标题（第1行，合并单元格）
    ws.merge_cells('A1:D1')
    title_cell = ws.cell(row=1, column=1, value='开发全周期记录')
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    title_cell.border = thin_border
    ws.row_dimensions[1].height = 30
    
    # 写入列标题（第2行）
    headers = ['时间', '版本', '更新标题', '更新内容摘要']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[2].height = 25
    
    # 更新日志数据
    changelog_data = [
        {
            'version': 'V0.9',
            'date': '2026-02-10',
            'title': '表格文本对齐优化',
            'content': [
                '• 统一模板编辑和报告编辑界面中表格内可编辑文本框的占位符文本对齐方式为居中对齐',
                '• 修复报告编辑界面中表格占位符文本未继承模板对齐设置的问题',
                '• 确保模板编辑界面的对齐样式修改能正确同步到报告编辑界面',
                '• 优化占位符文本样式继承机制，实现真正的模板-报告一致性',
                '• 修复 EditableField 组件中 placeholderFormat 对象定义语法错误'
            ]
        },
        {
            'version': 'V0.8',
            'date': '2026-02-06',
            'title': '稳定性与体验优化',
            'content': [
                '• 修复"测试条件"小标题样式不一致的问题',
                '• 统一各区域标题的显示效果和间距',
                '• 修复新增行无法自动保存的缺陷',
                '• 优化自动保存机制，确保空白行也能被正确保存',
                '• 改善用户体验，减少因刷新导致的数据丢失问题'
            ]
        },
        {
            'version': 'V0.7',
            'date': '2026-02-05',
            'title': '行操作界面优化',
            'content': [
                '• 重新设计行操作控件布局，将分散按钮整合到左侧功能侧边栏',
                '• 统一控件高度和间距，提升视觉一致性',
                '• 优化控件颜色策略：绿色(批量)、蓝色(指定)、橙色(警告)',
                '• 将"批量添加行"更名为"末尾添加行"，命名更加准确',
                '• 新增"指定位置插入行"功能，支持精确位置控制',
                '• 新增"末尾删除行"功能，可批量从末尾删除指定行数'
            ]
        },
        {
            'version': 'V0.6',
            'date': '2026-02-04',
            'title': '多页分页布局优化',
            'content': [
                '• 拆分"测试条件"区域为三个独立可分页元素',
                '• 修复测试图片行接触页面底部时被遮挡的问题',
                '• 修复内容减少后无法正确回流到原页面的问题',
                '• 修复第二页及以后页面的换页判定阈值与第一页不一致的问题',
                '• 实现动态行高测量机制，替代固定行高估算',
                '• 优化 DOM 测量机制，使用防抖监听内容变化'
            ]
        },
        {
            'version': 'V0.5',
            'date': '2026-02-02',
            'title': '报告编辑界面重构',
            'content': [
                '• 重构报告编辑界面布局，新增统一的左侧功能侧边栏',
                '• 将导出PDF和新建报告功能移至左侧边栏',
                '• 新增保存草稿功能：用户可为当前编辑内容命名并保存',
                '• 新增打开草稿功能：支持从已保存草稿列表中选择加载',
                '• 新增删除草稿功能：提供安全的草稿删除机制',
                '• 删除操作具有双重确认机制，防止误删重要数据'
            ]
        },
        {
            'version': 'V0.4',
            'date': '2026-02-02',
            'title': '模板应用机制重构',
            'content': [
                '• 重构模板与报告编辑隔离机制：模板编辑页面的修改不再自动同步',
                '• 新增"已应用模板"概念：只有点击"应用于报告编辑"按钮后才会应用',
                '• 优化"新建报告"功能：自动加载最后应用的模板状态',
                '• 修复模板编辑中可编辑字段样式无法正确应用到报告编辑的问题',
                '• 修复点击"新建报告"后模板样式重置的问题',
                '• 修复报告编辑工具栏格式修改功能不生效的问题'
            ]
        },
        {
            'version': 'V0.3',
            'date': '2026-01-29',
            'title': '样式持久化修复',
            'content': [
                '• 修复模板编辑界面中文本样式修改无法自动保存的问题',
                '• 修复修改某个文本框样式时导致其他同类型文本框被同步修改的问题',
                '• 修复页面导航后样式状态丢失的问题',
                '• 修复"应用于报告编辑"功能中样式修改无法成功应用的问题',
                '• 修复报告编辑界面页脚区域默认样式与模板编辑界面不一致的问题',
                '• 新增单独文本框样式持久化功能，每个文本框可独立保存样式'
            ]
        },
        {
            'version': 'V0.2',
            'date': '2026-01-23',
            'title': '功能优化与缺陷修复',
            'content': [
                '• 新增文本框对齐功能，支持左对齐、居中对齐、右对齐',
                '• 在模板编辑界面侧边栏新增模板操作区域',
                '• 修复模板编辑界面中图片移动时错位、超出 PDF 范围的问题',
                '• 修复文本框无法换行的问题',
                '• 修复报告编辑页面编辑文本样式时影响提示文本的问题',
                '• 调整框选触发方式为 Shift + 拖拽框选，避免误触发'
            ]
        },
        {
            'version': 'V0.1',
            'date': '2026-01-21',
            'title': '初始版本发布',
            'content': [
                '• 报告编辑器：支持通用模板的报告编辑功能',
                '• 格式工具栏：支持字体、字号、颜色、加粗、斜体等格式设置',
                '• 批量选择：支持 Ctrl + 点击多选进行批量格式设置',
                '• 动态行管理：测试结果和测试图片区域支持行的添加和删除',
                '• 图片上传：支持点击上传、拖拽上传和多图上传',
                '• PDF 导出：将报告内容导出为规范的 PDF 文档',
                '• 模板编辑器：支持自定义模板的固定内容样式',
                '• 自动保存：编辑内容自动保存，防止数据丢失'
            ]
        }
    ]
    
    # 写入数据
    current_row = 3
    for entry in changelog_data:
        # 合并更新内容为一个字符串
        content_text = '\n'.join(entry['content'])
        
        # 写入单元格
        ws.cell(row=current_row, column=1, value=entry['date'])
        ws.cell(row=current_row, column=2, value=entry['version'])
        ws.cell(row=current_row, column=3, value=entry['title'])
        ws.cell(row=current_row, column=4, value=content_text)
        
        # 应用样式
        ws.cell(row=current_row, column=1).font = content_font
        ws.cell(row=current_row, column=1).alignment = date_alignment
        ws.cell(row=current_row, column=1).border = thin_border
        
        ws.cell(row=current_row, column=2).font = content_font
        ws.cell(row=current_row, column=2).alignment = version_alignment
        ws.cell(row=current_row, column=2).border = thin_border
        
        ws.cell(row=current_row, column=3).font = content_font
        ws.cell(row=current_row, column=3).alignment = title_content_alignment
        ws.cell(row=current_row, column=3).border = thin_border
        
        ws.cell(row=current_row, column=4).font = content_font
        ws.cell(row=current_row, column=4).alignment = content_alignment
        ws.cell(row=current_row, column=4).border = thin_border
        
        # 设置行高（根据内容自动调整）
        ws.row_dimensions[current_row].height = max(60, len(entry['content']) * 15)
        
        current_row += 1
    
    # 冻结前两行
    ws.freeze_panes = 'A3'
    
    # 保存文件
    filename = '可靠性测试报告工具_更新日志汇总.xlsx'
    wb.save(filename)
    print(f'✓ Excel文件已生成：{filename}')
    return filename

if __name__ == '__main__':
    create_changelog_excel()
